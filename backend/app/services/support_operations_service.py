from collections import Counter
import csv
from io import BytesIO, StringIO
from pathlib import Path
import re
import subprocess
import tempfile
import uuid
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import AuditLog, KnowledgeChunk, KnowledgeDocument, SupportUnansweredQuestion
from app.services.knowledge_service import ingest_document


PROJECT_ROOT = Path(__file__).resolve().parents[3]
UPLOAD_ROOT = PROJECT_ROOT / "uploads" / "knowledge_sources"
SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
SUPPORTED_KNOWLEDGE_FILE_EXTENSIONS = {".csv", ".txt", ".md", ".markdown", ".pdf", ".docx", ".xlsx", *SUPPORTED_IMAGE_EXTENSIONS}


def record_unanswered_question(
    db: Session,
    *,
    question: str,
    user_id: str | None = None,
    customer_id: str | None = None,
    project_id: str | None = None,
    source: str = "kb_answer",
) -> SupportUnansweredQuestion | None:
    normalized_question = question.strip()
    if not normalized_question:
        return None

    existing = db.scalar(
        select(SupportUnansweredQuestion).where(
            SupportUnansweredQuestion.question == normalized_question,
            SupportUnansweredQuestion.status == "pending",
        )
    )
    if existing:
        return existing

    row = SupportUnansweredQuestion(
        question=normalized_question,
        user_id=user_id,
        customer_id=customer_id,
        project_id=project_id,
        source=source,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def list_unanswered_questions(db: Session, *, limit: int = 25, status: str | None = None) -> list[dict[str, Any]]:
    query = select(SupportUnansweredQuestion)
    if status:
        query = query.where(SupportUnansweredQuestion.status == status)
    rows = db.scalars(query.order_by(SupportUnansweredQuestion.created_at.desc()).limit(limit)).all()
    return [_serialize_unanswered(row) for row in rows]


def update_unanswered_question_status(db: Session, unanswered_id: str, *, status: str) -> dict[str, Any]:
    if status not in {"pending", "resolved", "ignored"}:
        raise ValueError("status must be pending, resolved, or ignored")
    row = db.get(SupportUnansweredQuestion, unanswered_id)
    if row is None:
        raise ValueError("unanswered question not found")
    row.status = status
    db.commit()
    db.refresh(row)
    return _serialize_unanswered(row)


def get_support_operations_dashboard(db: Session, *, limit: int = 10) -> dict[str, Any]:
    total_queries = _count_kb_answer_logs(db)
    unanswered_count = int(db.scalar(select(func.count()).select_from(SupportUnansweredQuestion)) or 0)
    missed_queries = max(_count_missed_answer_logs(db), unanswered_count)
    hit_queries = max(total_queries - missed_queries, 0)
    document_count = int(db.scalar(select(func.count()).select_from(KnowledgeDocument)) or 0)
    chunk_count = int(db.scalar(select(func.count()).select_from(KnowledgeChunk)) or 0)
    parsed_document_count = int(db.scalar(select(func.count()).select_from(KnowledgeDocument).where(KnowledgeDocument.parse_status == "parsed")) or 0)
    indexed_document_count = int(db.scalar(select(func.count()).select_from(KnowledgeDocument).where(KnowledgeDocument.index_status == "indexed")) or 0)

    return {
        "ok": True,
        "metrics": {
            "knowledge_documents": document_count,
            "answer_fragments": chunk_count,
            "parsed_documents": parsed_document_count,
            "indexed_documents": indexed_document_count,
            "total_queries": total_queries,
            "hit_queries": hit_queries,
            "missed_queries": missed_queries,
            "unanswered_questions": unanswered_count,
            "hit_rate": round(hit_queries / total_queries, 4) if total_queries else 0,
        },
        "popular_questions": _popular_questions(db, limit),
        "recent_unanswered": list_unanswered_questions(db, limit=limit),
        "recent_documents": _recent_documents(db, limit),
    }


def import_faq_text(
    db: Session,
    *,
    text: str,
    user_id: str,
    source_type: str = "faq_import",
    customer_id: str | None = None,
    project_id: str | None = None,
) -> dict[str, Any]:
    rows = _parse_faq_rows(text)
    imported = []
    for question, answer in rows:
        document = ingest_document(
            db,
            payload={
                "title": question,
                "summary": answer[:120],
                "content_text": f"问题：{question}\n答案：{answer}",
                "source_type": source_type,
                "customer_id": customer_id,
                "project_id": project_id,
            },
            user_id=user_id,
        )
        imported.append({"id": document.id, "title": document.title})

    return {
        "ok": True,
        "imported_count": len(imported),
        "documents": imported,
    }


def import_knowledge_file(
    db: Session,
    *,
    filename: str,
    content: bytes,
    user_id: str,
    source_type: str | None = None,
    customer_id: str | None = None,
    project_id: str | None = None,
) -> dict[str, Any]:
    safe_filename = _safe_filename(filename)
    extension = Path(safe_filename).suffix.lower()
    if extension not in SUPPORTED_KNOWLEDGE_FILE_EXTENSIONS:
        raise ValueError("unsupported file type; supported: csv, txt, md, markdown, pdf, docx, xlsx, png, jpg, jpeg, webp, bmp")
    if not content:
        raise ValueError("file is empty")

    try:
        parsed_text = _extract_file_text(extension, content)
    except Exception as error:
        raise ValueError(f"failed to parse file: {error}") from error
    stored_path = _store_uploaded_file(safe_filename, content)
    title = Path(safe_filename).stem if extension in SUPPORTED_IMAGE_EXTENSIONS else _derive_document_title(safe_filename, parsed_text)
    document = ingest_document(
        db,
        payload={
            "title": title,
            "summary": _build_summary(parsed_text),
            "content_text": parsed_text,
            "source_type": source_type or extension.lstrip(".") or "file_import",
            "source_file_path": str(stored_path),
            "source_file_name": safe_filename,
            "source_file_mime_type": _mime_type_for_extension(extension),
            "source_file_size": len(content),
            "source_file_storage": "uploaded_copy",
            "parse_status": "parsed",
            "index_status": "indexed",
            "customer_id": customer_id,
            "project_id": project_id,
        },
        user_id=user_id,
    )
    return {
        "ok": True,
        "imported_count": 1,
        "documents": [{"id": document.id, "title": document.title}],
        "source_file": {
            "name": document.source_file_name,
            "path": document.source_file_path,
            "storage": document.source_file_storage,
        },
    }


def _parse_faq_rows(text: str) -> list[tuple[str, str]]:
    stripped = text.strip()
    if not stripped:
        return []

    if "," in stripped.splitlines()[0]:
        reader = csv.DictReader(StringIO(stripped))
        if reader.fieldnames and {"问题", "答案"}.issubset(set(reader.fieldnames)):
            return [
                (str(row.get("问题") or "").strip(), str(row.get("答案") or "").strip())
                for row in reader
                if str(row.get("问题") or "").strip() and str(row.get("答案") or "").strip()
            ]

    if stripped.startswith("#"):
        lines = stripped.splitlines()
        title = lines[0].lstrip("#").strip()
        body = "\n".join(lines[1:]).strip()
        if title and body:
            return [(title, body)]

    rows: list[tuple[str, str]] = []
    for block in stripped.split("\n\n"):
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        question = ""
        answer = ""
        for line in lines:
            if line.startswith(("问题：", "问题:", "Q:", "Q：")):
                question = line.split(":", 1)[-1].split("：", 1)[-1].strip()
            elif line.startswith(("答案：", "答案:", "A:", "A：")):
                answer = line.split(":", 1)[-1].split("：", 1)[-1].strip()
        if question and answer:
            rows.append((question, answer))
    if rows:
        return rows

    plain_lines = [line.strip() for line in stripped.splitlines() if line.strip()]
    if len(plain_lines) >= 2:
        return [(plain_lines[0], "\n".join(plain_lines[1:]))]

    return []


def _safe_filename(filename: str) -> str:
    name = Path(filename).name.strip()
    if not name:
        return "uploaded-document"
    return re.sub(r"[\\/:*?\"<>|]+", "_", name)


def _store_uploaded_file(filename: str, content: bytes) -> Path:
    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    stored_path = UPLOAD_ROOT / f"{Path(filename).stem}---{uuid.uuid4()}{Path(filename).suffix}"
    stored_path.write_bytes(content)
    return stored_path


def _extract_file_text(extension: str, content: bytes) -> str:
    if extension in {".txt", ".md", ".markdown", ".csv"}:
        return content.decode("utf-8-sig")
    if extension in SUPPORTED_IMAGE_EXTENSIONS:
        return _extract_image_file_text(content)
    if extension == ".docx":
        return _extract_docx_text(content)
    if extension == ".xlsx":
        return _extract_xlsx_text(content)
    if extension == ".pdf":
        return _extract_pdf_text(content)
    raise ValueError("unsupported file type")


def _extract_docx_text(content: bytes) -> str:
    from docx import Document

    document = Document(BytesIO(content))
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    sections = []
    if paragraphs:
        sections.append("\n".join(paragraphs))
    image_sections = _extract_docx_image_sections(document, content, paragraphs)
    sections.extend(image_sections)
    text = "\n\n".join(sections)
    if not text.strip():
        raise ValueError("no readable text found in Word document")
    return text


def _extract_docx_image_sections(document: Any, content: bytes, paragraphs: list[str]) -> list[str]:
    image_blobs = _extract_docx_image_blobs(document, content)
    sections: list[str] = []
    for index, image_blob in enumerate(image_blobs, start=1):
        context = _docx_image_context(paragraphs, index, len(image_blobs))
        ocr_text = _extract_image_ocr_text(image_blob)
        lines = [f"### 文档图片 {index}"]
        if context:
            lines.append(f"图片上下文：{context}")
        if ocr_text:
            lines.append(f"图片OCR识别文字：{ocr_text}")
        else:
            lines.append("图片OCR识别文字：未识别到可读文字。")
        sections.append("\n".join(lines))
    return sections


def _extract_docx_image_blobs(document: Any, content: bytes) -> list[bytes]:
    blobs: list[bytes] = []
    seen: set[str] = set()
    for relationship in document.part.rels.values():
        if "image" not in relationship.reltype:
            continue
        blob = relationship.target_part.blob
        marker = str(hash(blob))
        if marker not in seen:
            blobs.append(blob)
            seen.add(marker)
    if blobs:
        return blobs

    from zipfile import ZipFile

    with ZipFile(BytesIO(content)) as archive:
        for name in archive.namelist():
            if name.startswith("word/media/"):
                blobs.append(archive.read(name))
    return blobs


def _docx_image_context(paragraphs: list[str], image_index: int, image_count: int) -> str:
    if not paragraphs:
        return ""
    if image_count <= 1:
        if len(paragraphs) == 1:
            return paragraphs[0]
        return f"{paragraphs[-2]} / {paragraphs[-1]}"
    anchor = min(len(paragraphs) - 1, max(0, round((image_index - 0.5) * len(paragraphs) / image_count)))
    start = max(0, anchor - 1)
    end = min(len(paragraphs), anchor + 2)
    return " / ".join(paragraphs[start:end])


def _extract_xlsx_text(content: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                rows.append("\t".join(values))
        if rows:
            sections.append(f"### {sheet.title}\n" + "\n".join(rows))
    text = "\n\n".join(sections).strip()
    if not text:
        raise ValueError("no readable text found in Excel document")
    return text


def _extract_pdf_text(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(content))
    pages = [(page.extract_text() or "").strip() for page in reader.pages]
    text = "\n\n".join(page for page in pages if page).strip()
    if not text:
        raise ValueError("no readable text found in PDF document")
    return text


def _extract_image_file_text(content: bytes) -> str:
    ocr_text = _extract_image_ocr_text(content)
    if ocr_text:
        return f"图片OCR识别文字：{ocr_text}"
    return "图片原件已保存。图片OCR识别文字：未识别到可读文字。"


def _extract_image_ocr_text(content: bytes) -> str:
    with tempfile.NamedTemporaryFile(suffix=".img") as image_file:
        image_file.write(content)
        image_file.flush()
        try:
            result = subprocess.run(
                ["tesseract", image_file.name, "stdout", "-l", "chi_sim+eng"],
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
        except (OSError, subprocess.SubprocessError):
            return ""
    if result.returncode != 0:
        return ""
    text = result.stdout
    return "\n".join(line.strip() for line in text.splitlines() if line.strip())


def _derive_document_title(filename: str, text: str) -> str:
    first_line = next((line.strip().lstrip("#").strip() for line in text.splitlines() if line.strip()), "")
    return first_line[:120] if first_line else Path(filename).stem


def _build_summary(text: str) -> str:
    compact = " ".join(line.strip() for line in text.splitlines() if line.strip())
    return compact[:160]


def _mime_type_for_extension(extension: str) -> str:
    return {
        ".csv": "text/csv",
        ".txt": "text/plain",
        ".md": "text/markdown",
        ".markdown": "text/markdown",
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
    }.get(extension, "application/octet-stream")


def _count_kb_answer_logs(db: Session) -> int:
    return int(db.scalar(select(func.count()).select_from(AuditLog).where(AuditLog.tool_name == "kb_answer")) or 0)


def _count_missed_answer_logs(db: Session) -> int:
    return int(
        db.scalar(
            select(func.count())
            .select_from(AuditLog)
            .where(AuditLog.tool_name == "kb_answer", AuditLog.response_summary.contains("没有找到"))
        )
        or 0
    )


def _popular_questions(db: Session, limit: int) -> list[dict[str, Any]]:
    logs = db.scalars(select(AuditLog).where(AuditLog.tool_name == "kb_answer")).all()
    counts: Counter[str] = Counter()
    for log in logs:
        question = _extract_query(log.request_payload)
        if question:
            counts[question] += 1
    return [{"question": question, "count": count} for question, count in counts.most_common(limit)]


def _extract_query(payload: dict[str, Any]) -> str:
    input_payload = payload.get("input") if isinstance(payload.get("input"), dict) else payload
    value = input_payload.get("query") if isinstance(input_payload, dict) else None
    return str(value or "").strip()


def _recent_documents(db: Session, limit: int) -> list[dict[str, Any]]:
    rows = db.scalars(select(KnowledgeDocument).order_by(KnowledgeDocument.updated_at.desc()).limit(limit)).all()
    return [
        {
            "id": row.id,
            "title": row.title,
            "summary": row.summary,
            "source_type": row.source_type,
            "parse_status": row.parse_status,
            "index_status": row.index_status,
            "source_file_name": row.source_file_name,
            "source_channel": row.source_channel,
            "source_external_user_id": row.source_external_user_id,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        }
        for row in rows
    ]


def _serialize_unanswered(row: SupportUnansweredQuestion) -> dict[str, Any]:
    return {
        "id": row.id,
        "question": row.question,
        "customer_id": row.customer_id,
        "project_id": row.project_id,
        "user_id": row.user_id,
        "status": row.status,
        "source": row.source,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }
